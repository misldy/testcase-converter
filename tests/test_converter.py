import unittest
import tempfile
import os
from pathlib import Path
from testcase_converter.converter import TestCaseConverter, ConversionType, TestCase

class TestCaseConverterTest(unittest.TestCase):
    def setUp(self):
        # 创建临时目录
        self.temp_dir = tempfile.mkdtemp()
        self.log_dir = Path(self.temp_dir) / "logs"
        self.log_dir.mkdir()
        
        # 测试数据
        self.sample_case = TestCase(
            module_path="模块→子模块",
            name="测试用例1",
            precondition="前置条件",
            steps="步骤1\n步骤2",
            expected_result="预期结果",
            vehicle_type="车型A",
            priority="1"
        )

    def tearDown(self):
        # 先关闭所有日志处理器
        import logging
        for handler in logging.getLogger('testcase_converter').handlers[:]:
            handler.close()
            logging.getLogger('testcase_converter').removeHandler(handler)
        
        # 清理临时目录
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_detect_conversion_type(self):
        # 测试转换类型检测
        with tempfile.NamedTemporaryFile(suffix='.xlsx', dir=self.temp_dir) as tmp:
            converter = TestCaseConverter(tmp.name)
            self.assertEqual(converter.conversion_type, ConversionType.EXCEL_TO_XMIND)
        
        with tempfile.NamedTemporaryFile(suffix='.xmind', dir=self.temp_dir) as tmp:
            converter = TestCaseConverter(tmp.name)
            self.assertEqual(converter.conversion_type, ConversionType.XMIND_TO_EXCEL)

    def test_parse_excel_row(self):
        # 测试Excel行解析
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["模块→子模块", "测试用例1", "前置条件", "步骤1\n步骤2", "预期结果", "车型A", "1"])
        
        excel_path = os.path.join(self.temp_dir, "test.xlsx")
        wb.save(excel_path)
        
        converter = TestCaseConverter(excel_path)
        test_case = converter._parse_excel_row([cell.value for cell in ws[1]])
        self.assertEqual(test_case.module_path, self.sample_case.module_path)
        self.assertEqual(test_case.name, self.sample_case.name)

    def test_excel_to_xmind_conversion(self):
        # 测试Excel转XMind
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "测试工作表"
        ws.append(["模块", "用例名称", "前置条件", "执行步骤", "预期结果", "车型", "优先级"])
        ws.append([self.sample_case.module_path, self.sample_case.name, 
                  self.sample_case.precondition, self.sample_case.steps, 
                  self.sample_case.expected_result, self.sample_case.vehicle_type, 
                  self.sample_case.priority])
        
        excel_path = os.path.join(self.temp_dir, "test.xlsx")
        wb.save(excel_path)
        
        converter = TestCaseConverter(excel_path)
        converter.convert()
        
        # 验证XMind文件是否生成
        xmind_files = list(Path(self.temp_dir).glob("*.xmind"))
        self.assertTrue(len(xmind_files) > 0)

    def test_xmind_to_excel_conversion(self):
        # 测试XMind转Excel - 创建真实XMind文件结构
        import xmind
        xmind_workbook = xmind.load("template.xmind")
        sheet = xmind_workbook.getPrimarySheet()
        root_topic = sheet.getRootTopic()
        root_topic.setTitle("根节点")
        
        # 添加测试用例
        test_topic = root_topic.addSubTopic()
        test_topic.setTitle(self.sample_case.name)
        test_topic.setPlainNotes(
            f"【前置条件】{self.sample_case.precondition}\n"
            f"【执行步骤】{self.sample_case.steps}\n"
            f"【预期结果】{self.sample_case.expected_result}\n"
            f"【车型】{self.sample_case.vehicle_type}\n"
            f"【优先级】{self.sample_case.priority}"
        )
        
        xmind_path = os.path.join(self.temp_dir, "test.xmind")
        xmind.save(xmind_workbook, xmind_path)
        
        converter = TestCaseConverter(xmind_path)
        converter.convert()
        
        # 验证Excel文件是否生成且包含数据
        excel_files = list(Path(self.temp_dir).glob("*.xlsx"))
        self.assertTrue(len(excel_files) > 0)
        
        # 验证Excel内容
        from openpyxl import load_workbook
        wb = load_workbook(excel_files[0])
        self.assertTrue(len(wb.sheetnames) > 0)
        ws = wb.active
        self.assertGreaterEqual(ws.max_row, 2)  # 至少包含表头和数据行

if __name__ == "__main__":
    unittest.main()