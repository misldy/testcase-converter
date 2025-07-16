"""
测试用例格式转换工具

此模块提供了Excel和XMind格式测试用例之间的双向转换功能。
支持：
1. Excel -> XMind 转换
2. XMind -> Excel 转换
"""

import xmind
import pkg_resources
from xmind.core.topic import TopicElement
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import uuid
import logging
from typing import List, Dict, Optional, Tuple, Any
from pathlib import Path
from dataclasses import dataclass
from enum import Enum

class ConversionType(Enum):
    """转换类型枚举"""
    EXCEL_TO_XMIND = "excel_to_xmind"
    XMIND_TO_EXCEL = "xmind_to_excel"

@dataclass
class TestCase:
    """测试用例数据结构"""
    module_path: str
    name: str
    precondition: Optional[str]
    steps: Optional[str]
    expected_result: Optional[str]
    vehicle_type: str
    priority: str

class ConverterConfig:
    """转换器配置类"""
    # 优先级有效值
    PRIORITY_VALUES = {str(i) for i in range(6)} | {i for i in range(6)}
    
    # Excel文件列定义
    COLUMN_NAMES = {
        'MODULE': '模块',
        'CASE_NAME': '用例名称',
        'PRECONDITION': '前置条件',
        'STEPS': '执行步骤',
        'EXPECTED': '预期结果',
        'VEHICLE_TYPE': '车型',
        'PRIORITY': '优先级'
    }

    # Excel样式配置
    COLUMN_WIDTHS = {
        'A': 30,  # 模块
        'B': 30,  # 用例名称
        'C': 45,  # 前置条件
        'D': 45,  # 执行步骤
        'E': 45,  # 预期结果
        'F': 20,  # 车型
        'G': 10   # 优先级
    }
    
    HEADER_STYLE = {
        'font': Font(name='Calibri', size=16, bold=True),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'fill': PatternFill("solid", fgColor="D3D3D3")
    }
    
    CELL_BORDER = Border(
        left=Side(border_style='thin', color='D3D3D3'),
        right=Side(border_style='thin', color='D3D3D3'),
        top=Side(border_style='thin', color='D3D3D3'),
        bottom=Side(border_style='thin', color='D3D3D3')
    )
    
    # 日志格式
    LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'
    LOG_FILENAME = 'conversion.log'

class TestCaseConverter:
    """测试用例格式转换器"""
    
    def _detect_conversion_type(self, input_file: str) -> ConversionType:
        """
        自动检测文件类型并返回对应的转换类型
        
        Args:
            input_file: 输入文件路径
        
        Returns:
            ConversionType: 转换类型
        
        Raises:
            ValueError: 如果文件扩展名不是 .xlsx 或 .xmind
        """
        file_ext = Path(input_file).suffix.lower()
        if file_ext == '.xlsx':
            return ConversionType.EXCEL_TO_XMIND
        elif file_ext == '.xmind':
            return ConversionType.XMIND_TO_EXCEL
        else:
            raise ValueError(f"不支持的文件类型: {file_ext}. 仅支持 .xlsx 和 .xmind 文件")

    def __init__(self, input_file: str, conversion_type: Optional[ConversionType] = None):
        """
        初始化转换器
        
        Args:
            input_file: 输入文件路径
            conversion_type: 可选的转换类型，如果未提供则自动检测
        """
        self.input_path = Path(input_file)
        if not self.input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
            
        # 如果未指定转换类型，则自动检测
        self.conversion_type = conversion_type or self._detect_conversion_type(input_file)
        self.file_name = self.input_path.stem
        self.output_directory = self.input_path.parent
        self._setup_logging()

    def _setup_logging(self) -> None:
        """配置日志记录器"""
        log_file = self.output_directory / ConverterConfig.LOG_FILENAME
        logging.basicConfig(
            level=logging.INFO,
            format=ConverterConfig.LOG_FORMAT,
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler(log_file, encoding='utf-8')
            ]
        )
        logging.info(f"开始处理文件: {self.input_path}")

    # ====== Excel转XMind相关方法 ======

    def _unpack_merged_cells(self, worksheet) -> None:
        """解开合并的单元格"""
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_cell in merged_ranges:
            top_left_value = worksheet.cell(
                row=merged_cell.min_row, 
                column=merged_cell.min_col
            ).value
            worksheet.unmerge_cells(str(merged_cell))
            
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    worksheet.cell(row=row, column=col).value = top_left_value

    def _parse_excel_row(self, row_data: Tuple[Any, ...]) -> Optional[TestCase]:
        """解析Excel行数据"""
        try:
            module_path, name, precondition, steps, expected, vehicle_type, priority = row_data
            return TestCase(
                module_path=str(module_path),
                name=str(name),
                precondition=str(precondition) if precondition else None,
                steps=str(steps) if steps else None,
                expected_result=str(expected) if expected else None,
                vehicle_type=str(vehicle_type),
                priority=str(priority)
            )
        except (ValueError, TypeError):
            return None

    def _validate_test_case(self, test_case: TestCase, sheet_name: str, row_index: int) -> bool:
        """验证测试用例数据"""
        if not test_case.module_path or not test_case.name or not test_case.vehicle_type:
            logging.warning(f"工作表 '{sheet_name}' 第 {row_index} 行模块/用例名称/车型为空")
            return False

        if test_case.priority not in ConverterConfig.PRIORITY_VALUES:
            logging.warning(f"工作表 '{sheet_name}' 第 {row_index} 行优先级无效: {test_case.priority}")
            return False

        return True

    def _create_xmind_notes(self, test_case: TestCase) -> str:
        """创建XMind节点的备注内容"""
        notes_parts = [
            ('前置条件', test_case.precondition),
            ('执行步骤', test_case.steps),
            ('预期结果', test_case.expected_result),
            ('车型', test_case.vehicle_type),
        ]
        notes = '\n'.join(
            f"【{name}】\n{value}" 
            for name, value in notes_parts 
            if value
        )
        return f"{notes}\n【优先级】{test_case.priority}" if notes else f"【优先级】{test_case.priority}"

    def _get_or_create_subtopic(self, parent_topic: TopicElement, title: str, workbook) -> TopicElement:
        """获取或创建子主题"""
        for subtopic in parent_topic.getSubTopics():
            if subtopic.getTitle() == title:
                return subtopic
        
        new_topic = TopicElement(ownerWorkbook=workbook)
        new_topic.setTitle(title)
        parent_topic.addSubTopic(new_topic)
        return new_topic

    def _create_topic_hierarchy(self, workbook, root_topic: TopicElement, test_case: TestCase) -> None:
        """创建主题层级结构"""
        current_topic = root_topic
        
        for module in test_case.module_path.split('→'):
            current_topic = self._get_or_create_subtopic(current_topic, module, workbook)

        case_topic = self._get_or_create_subtopic(current_topic, test_case.name, workbook)
        case_topic.setPlainNotes(self._create_xmind_notes(test_case))

    def _excel_to_xmind(self) -> None:
        """执行Excel到XMind的转换"""
        workbook = load_workbook(filename=self.input_path)
        logging.info(f"成功加载Excel文件，共包含 {len(workbook.worksheets)} 个工作表")
        
        for sheet in workbook.worksheets:
            try:
                self._process_excel_sheet(sheet)
            except Exception as e:
                logging.error(f"处理工作表 '{sheet.title}' 时出错: {e}")
                continue
                
        logging.info("转换完成")

    def _process_excel_sheet(self, sheet) -> None:
        """处理Excel工作表"""
        logging.info(f"开始处理工作表: {sheet.title}")
        self._unpack_merged_cells(sheet)

        template_path = pkg_resources.resource_filename(
            'testcase_converter', 'resources/template.xmind'
        )
        xmind_workbook = xmind.load(template_path)
        sheet_xmind = xmind_workbook.getPrimarySheet()
        if sheet_xmind is None:
            logging.error("未找到主工作表，请检查 template.xmind 文件。")
            return
        sheet_xmind.setTitle(sheet.title)
        root_topic = sheet_xmind.getRootTopic()
        root_topic.setTitle(sheet.title)

        processed_count = 0
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            test_case = self._parse_excel_row(row)
            if not test_case or not self._validate_test_case(test_case, sheet.title, index):
                continue
                
            self._create_topic_hierarchy(xmind_workbook, root_topic, test_case)
            processed_count += 1

        output_path = self.output_directory / f"{self.file_name}_{uuid.uuid4()}_{sheet.title}.xmind"
        try:
            xmind.save(xmind_workbook, path=str(output_path))
            logging.info(f"工作表 '{sheet.title}' 处理完成，成功转换 {processed_count} 条用例")
            logging.info(f"已保存XMind文件: {output_path}")
        except Exception as e:
            logging.error(f"保存XMind文件时出错: {e}")

    # ====== XMind转Excel相关方法 ======

    def _parse_xmind_notes(self, notes: Optional[str]) -> List[str]:
        """解析XMind备注内容"""
        if not notes:
            return [''] * 5

        notes_dict = {}
        for part in notes.split('【'):
            if not part.strip():
                continue
            try:
                key, value = part.split('】', 1)
                notes_dict[key] = value.strip()
            except ValueError:
                logging.warning(f"无效的备注格式: {part}")
                continue

        return [
            notes_dict.get('前置条件', ''),
            notes_dict.get('执行步骤', ''),
            notes_dict.get('预期结果', ''),
            notes_dict.get('车型', ''),
            notes_dict.get('优先级', '')
        ]

    def _process_xmind_topics(self, topics, parent_title: str = '') -> List[List[str]]:
        """递归处理XMind主题"""
        rows = []
        for topic in topics:
            title = topic.getTitle()
            if not title:
                continue

            full_title = f"{parent_title}→{title}" if parent_title else title
            subtopics = topic.getSubTopics()

            if subtopics:
                rows.extend(self._process_xmind_topics(subtopics, full_title))
            else:
                notes = topic.getNotes()
                notes_formatted = self._parse_xmind_notes(notes)
                try:
                    modules, case_name = full_title.rsplit('→', 1)
                except ValueError:
                    modules, case_name = '', full_title
                rows.append([modules, case_name] + notes_formatted)

        return rows

    def _sanitize_sheet_title(self, title: str) -> str:
        """
        清理工作表标题，移除Excel不允许的特殊字符
        
        Args:
            title: 原始标题
            
        Returns:
            str: 清理后的标题
        """
        # Excel工作表名称的限制：
        # 1. 不能包含以下字符: / \ [ ] * ? :
        # 2. 长度不能超过31个字符
        invalid_chars = ['/', '\\', '[', ']', '*', '?', ':']
        result = title
        for char in invalid_chars:
            result = result.replace(char, '_')
            
        # 截断到31个字符
        result = result[:31]
        
        # 确保不为空
        if not result.strip():
            result = f"Sheet_{uuid.uuid4().hex[:8]}"
            
        return result

    def _create_excel_worksheet(self, sheet_title: str, rows: List[List[str]]) -> Worksheet:
        """创建并格式化Excel工作表"""
        sanitized_title = self._sanitize_sheet_title(sheet_title)
        ws = self.excel_wb.create_sheet(title=sanitized_title)
        
        # 设置列宽
        for col, width in ConverterConfig.COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # 添加表头
        headers = list(ConverterConfig.COLUMN_NAMES.values())
        for col_num, title in enumerate(headers, 1):
            cell = ws[f'{get_column_letter(col_num)}1']
            cell.value = title
            for key, value in ConverterConfig.HEADER_STYLE.items():
                setattr(cell, key, value)

        # 添加数据
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, content in enumerate(row, 1):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.value = content
                cell.border = ConverterConfig.CELL_BORDER
                cell.alignment = Alignment(wrap_text=True)

        ws.sheet_view.showGridLines = False
        return ws

    def _xmind_to_excel(self) -> None:
        """执行XMind到Excel的转换"""
        try:
            xmind_wb = xmind.load(self.input_path)
            self.excel_wb = Workbook()
            if self.excel_wb.active is not None:
                self.excel_wb.remove(self.excel_wb.active)
            
            total_sheets = len(xmind_wb.getSheets())
            for i, sheet in enumerate(xmind_wb.getSheets(), 1):
                logging.info(f'处理工作表 {i}/{total_sheets}: {sheet.getTitle()}')
                
                root_topic = sheet.getRootTopic()
                topics = root_topic.getSubTopics()
                rows = self._process_xmind_topics(topics)
                
                self._create_excel_worksheet(root_topic.getTitle() or '', rows)

            output_path = self.output_directory / f"{self.file_name}_{uuid.uuid4()}.xlsx"
            self.excel_wb.save(filename=output_path)
            logging.info(f'转换完成，已保存Excel文件: {output_path}')
            
        except Exception as e:
            logging.error(f"转换过程中出错: {e}")
            raise

    def convert(self) -> None:
        """执行转换"""
        try:
            if self.conversion_type == ConversionType.EXCEL_TO_XMIND:
                self._excel_to_xmind()
            else:
                self._xmind_to_excel()
        except Exception as e:
            logging.error(f"转换失败: {e}")
            raise

def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='测试用例格式转换工具')
    parser.add_argument('input_file', help='输入文件路径')
    
    args = parser.parse_args()
    
    try:
        converter = TestCaseConverter(
            args.input_file
        )
        converter.convert()
    except Exception as e:
        logging.error(f"程序执行出错: {e}")
        raise

if __name__ == "__main__":
    main()